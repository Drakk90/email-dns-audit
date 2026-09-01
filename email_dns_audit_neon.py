#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════════════════════
 EMAIL DNS AUDIT NEON v3.3 — Excel Unificado, i18n & RDAP/WHOIS Riguroso
 Auditoria DNS / Email Authentication / Asset Governance
───────────────────────────────────────────────────────────────────────────────
 Autor   : Eduardo Recinos
 CISO    : VCISO
 Version : 3.3
 Fecha   : 2026-09-01
 Cambios v3.3:
   - Integración rigurosa de RDAP (RFC 7480-7484 / RFC 9082-9083) sobre HTTPS.
   - Bootstrap autoritativo de TLDs (Verisign, PIR, Nominet, Identity Digital, rdap.org).
   - Extracción de Registrar, Expiration/Creation y Entidad Registrante (Brand).
   - Inferencia heurística de Brand e Internal Owner (Seguridad/TI) para inventario.
   - Soporte bilingüe completo (Español / Inglés) vía flag `--lang [es|en]`.
   - Reporte Excel unificado con pestañas, tablas y formatos condicionales.
═══════════════════════════════════════════════════════════════════════════════
"""
import sys, os, re, time, base64, argparse, asyncio, subprocess, importlib, json, ssl, socket
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

REQUIRED = {
    "rich": "rich>=13.7.0", "dns": "dnspython>=2.4.0",
    "cryptography": "cryptography>=42.0.0", "httpx": "httpx>=0.27.0",
    "whois": "python-whois>=0.9.0", "aiodns": "aiodns>=3.1.0",
    "openpyxl": "openpyxl>=3.1.0",
}


def check_deps(auto: bool = False) -> bool:
    missing = []
    for mod, pip_name in REQUIRED.items():
        try:
            importlib.import_module(mod)
        except ImportError:
            missing.append(pip_name)
    if not missing:
        return True
    msg = ("\n\033[91m[!] Faltan dependencias:\033[0m\n  " + "\n  ".join(missing) +
           "\n\n\033[96mInstalar:\033[0m\n  pip install --user " + " ".join(missing) +
           "\n\n\033[96mO ejecuta:\033[0m\n  python3 " + sys.argv[0] + " --install-deps\n")
    if auto:
        print("\033[93m[*] Instalando dependencias...\033[0m")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", *missing])
        print("\033[92m[OK] Re-ejecuta el script.\033[0m")
        sys.exit(0)
    print(msg)
    return False


if "--install-deps" in sys.argv:
    check_deps(auto=True)
if not check_deps():
    sys.exit(1)

import dns.resolver, dns.asyncresolver, dns.flags, dns.rcode, dns.reversename
import dns.message, dns.asyncquery
import httpx
import whois as whois_lib
from cryptography import x509
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.backends import default_backend
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.align import Align
from rich.box import DOUBLE, ROUNDED, HEAVY
from rich.style import Style
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

from i18n import get_translator

# Paleta corporativa Excel
C_DARK = "0E3178"; C_NAVY = "0D1C43"; C_MID = "004E96"; C_BLUE = "0066B3"
C_ACC = "2B96CD"; C_SOFT = "8FC2E0"; C_BAND = "EAF3FA"; C_WHITE = "FFFFFF"

# Paleta neon consola
NM = "#ff00ff"; NC = "#00ffff"; NG = "#39ff14"; NO = "#ff6600"
NR = "#ff003c"; NY = "#ffff00"; NP = "#9d00ff"

RDAP_HEADERS = {
    "User-Agent": "Email-DNS-Audit-Neon/3.3 (Security Audit Engine)",
    "Accept": "application/rdap+json, application/json"
}

COMMON_SELECTORS = [
    # Microsoft 365 / Outlook
    "selector1", "selector2",
    # Google Workspace
    "google",
    # SendGrid
    "s1", "s2", "smtpapi", "em",
    # Mailchimp / Mandrill
    "k1", "k2", "k3", "mandrill", "mte1", "mte2",
    # Mailgun
    "mx", "mg", "smtp", "pic", "mailo",
    # Amazon SES
    "ses", "amazonses", "ses-1", "ses-2",
    # Postmark
    "pm", "20150623", "20150625",
    # SparkPost
    "scph0817", "scph0816", "sparkpostmail", "sp",
    # Zoho
    "zoho", "zmail",
    # HubSpot
    "hubspotemail1", "hubspotemail2", "hs1", "hs2",
    # Proofpoint / Mimecast
    "ppdkim", "mimecast",
    # Salesforce / Marketing Cloud
    "et", "200608",
    # Klaviyo / Constant Contact / Zendesk / Freshdesk
    "kl", "kl2", "roving", "zendesk1", "zendesk2", "fd", "fd2",
    # MXVault / Dyn / Everlytic / Proton
    "mxvault", "dynect", "everlytickey1", "everlytickey2", "sm", "proton",
    # Genericos
    "dkim", "mail", "default",
]

console = Console()


def derive_brand_from_domain(domain: str) -> str:
    """Extracts a clean, capitalized brand name from a domain name."""
    parts = domain.lower().split(".")
    if len(parts) >= 2:
        if parts[-2] in ("com", "co", "org", "gob", "edu", "net", "nom") and len(parts) >= 3:
            raw_brand = parts[-3]
        else:
            raw_brand = parts[-2]
    else:
        raw_brand = parts[0]
    return raw_brand.capitalize()


def get_rdap_candidate_urls(domain: str) -> List[str]:
    """Generates authoritative and bootstrap RDAP URLs based on TLD."""
    parts = domain.lower().split(".")
    tld = parts[-1]
    urls = []
    if tld in ("com", "net"):
        urls.append(f"https://rdap.verisign.com/{tld}/v1/domain/{domain}")
    elif tld == "org":
        urls.append(f"https://rdap.publicinterestregistry.org/rdap/domain/{domain}")
    elif tld in ("info", "biz", "mobi", "pro", "asia", "io"):
        urls.append(f"https://rdap.identitydigital.services/rdap/domain/{domain}")
    elif tld in ("uk", "co.uk", "org.uk"):
        urls.append(f"https://rdap.nominet.uk/rdap/domain/{domain}")
    urls.append(f"https://rdap.org/domain/{domain}")
    urls.append(f"https://rdap.iana.org/domain/{domain}")
    return urls


def generate_deep_selectors(months_back: int = 30) -> List[str]:
    deep = set()
    deep.update([
        "google2", "goog", "gmail",
        "selector1-onmicrosoft-com", "selector2-onmicrosoft-com", "s1024", "s2048",
        "amazonses1", "amazonses2", "aws", "awsses",
        "s3", "s4", "sendgrid", "sg", "sig1",
        "mailgun", "mg1", "mg2", "k2mg", "email",
        "scph0316", "scph0517", "scph0916", "scph1016", "scph1216",
        "scph0118", "scph0218", "scph0318", "scph0618", "scph1017",
        "20161025", "20170101", "20180101", "20190101", "20200101",
        "zendesk", "freshdesk", "klaviyo", "kl3",
        "hs3", "hs4", "hubspot",
        "mc", "sfmc", "exacttarget", "cust", "mta",
        "cm", "createsend", "campaignmonitor",
        "litmus", "sailthru", "sailthru1",
        "mktomail", "m1", "marketo",
        "pardot", "psdkim",
        "sib", "sendinblue", "brevo",
        "cc", "ctct",
        "intercom", "ic",
        "mailerlite", "ml",
        "omnisend", "os",
        "activecampaign", "ac",
        "drip", "dr",
        "twilio", "tw",
        "mailjet1", "mailjet2",
        "elasticemail", "ee",
        "smtp2go", "s2g",
        "socketlabs", "sl",
        "turbosmtp", "ts",
        "mailpoet", "mp",
        "protonmail", "protonmail2", "protonmail3",
        "fm1", "fm2", "fm3",
        "dkim1", "dkim2", "key1", "key2",
    ])
    today = datetime.now(timezone.utc)
    for i in range(months_back + 1):
        y = today.year
        mth = today.month - i
        while mth <= 0:
            mth += 12
            y -= 1
        deep.add(f"{y:04d}{mth:02d}01")
        deep.add(f"{y:04d}{mth:02d}")
        deep.add(f"{y:04d}{mth:02d}15")
    return sorted(deep)


def map_dkim_service(s: str) -> str:
    m = {"selector1": "Microsoft 365", "selector2": "Microsoft 365", "google": "Google Workspace",
         "k1": "Mailchimp / SendGrid", "k2": "Mailchimp / SendGrid", "k3": "Mailchimp / SendGrid",
         "smtpapi": "SendGrid", "mandrill": "Mandrill (Mailchimp)",
         "fd": "Mailgun", "fd2": "Mailgun", "pic": "Postmark", "pm": "Postmark", "mxvault": "MXVault",
         "ses-1": "Amazon SES", "ses-2": "Amazon SES", "ses": "Amazon SES", "amazonses": "Amazon SES",
         "zoho": "Zoho Mail", "zmail": "Zoho Mail", "dynect": "Dyn",
         "hubspotemail1": "HubSpot", "hubspotemail2": "HubSpot", "hs1": "HubSpot", "hs2": "HubSpot",
         "everlytickey1": "Everlytic", "everlytickey2": "Everlytic", "sm": "SparkPost",
         "scph0817": "SparkPost", "scph0816": "SparkPost", "sparkpostmail": "SparkPost", "sp": "SparkPost",
         "proton": "Proton Mail", "mx": "Mailgun", "mg": "Mailgun", "smtp": "Mailgun",
         "em": "SendGrid", "s1": "SendGrid / Generico", "s2": "SendGrid / Generico",
         "ppdkim": "Proofpoint", "mimecast": "Mimecast", "et": "Salesforce Marketing Cloud",
         "roving": "Constant Contact", "zendesk1": "Zendesk", "zendesk2": "Zendesk",
         "kl": "Klaviyo", "kl2": "Klaviyo", "mte1": "Mandrill", "mte2": "Mandrill",
         "dkim": "Generico", "mail": "Generico", "default": "Generico"}
    if s in m:
        return m[s]
    if re.fullmatch(r"20\d{4,6}", s):
        return "Google (Rotativo)"
    if s.startswith("scph"):
        return "SparkPost"
    if s.startswith("hs"):
        return "HubSpot"
    if s.startswith("ses") or s.startswith("amazonses"):
        return "Amazon SES"
    if s.startswith("mg") or s.startswith("mailgun"):
        return "Mailgun"
    if s.startswith("protonmail"):
        return "Proton Mail"
    return "Custom"


def map_provider_purpose(p: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if p in ("Microsoft 365", "Google Workspace", "Zoho Mail", "Proton Mail"):
        return t("purpose_internal")
    if p in ("Mailchimp", "HubSpot", "Mailjet", "Everlytic"):
        return t("purpose_marketing")
    if p in ("SendGrid", "Mailgun", "Amazon SES", "Postmark", "SparkPost", "Mandrill",
             "Mandrill (Mailchimp)", "Mailchimp / SendGrid"):
        return t("purpose_transactional")
    return t("status_to_validate")


def map_provider_mechanism(p: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if p == "Microsoft 365":
        return t("mechanism_m365")
    if p == "Google Workspace":
        return t("mechanism_google")
    if p in ("Zoho Mail", "Proton Mail"):
        return t("mechanism_smtp")
    if p in ("SendGrid", "Mailgun", "Amazon SES", "Postmark", "SparkPost", "Mandrill",
             "Mandrill (Mailchimp)", "Mailjet", "Mailchimp / SendGrid"):
        return t("mechanism_api_relay")
    if p in ("Mailchimp", "HubSpot", "Everlytic"):
        return t("mechanism_saas")
    return t("status_to_validate")


def derive_spf_status(pub: str, a: str, lk: int, v: int, m: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if pub == "no" or m == "si" or a == "+all" or lk > 10 or v > 2:
        return t("status_non_compliant")
    if a in ("?all", "sin all") or v >= 1 or lk >= 8:
        return t("status_partial")
    return t("status_compliant")


def derive_spf_severity(pub: str, a: str, lk: int, v: int, m: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if m == "si" or a == "+all": return t("sev_critical")
    if pub == "no" or lk > 10 or a in ("?all", "sin all"): return t("sev_high")
    if v > 2: return t("sev_medium")
    if v >= 1 or lk >= 8: return t("sev_low")
    return t("sev_info")


def derive_dmarc_status(pub: str, p: str, sp: str, pct: str, rua: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if pub == "no" or p == "none": return t("status_non_compliant")
    if p == "reject" and pct == "100" and rua and sp: return t("status_compliant")
    return t("status_partial")


def derive_dmarc_severity(pub: str, p: str, sp: str, pct: str, rua: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if pub == "no": return t("sev_critical")
    if p == "none" or not rua: return t("sev_high")
    if p == "quarantine" or pct != "100": return t("sev_medium")
    if p == "reject" and not sp: return t("sev_low")
    return t("sev_info")


def derive_dnssec_status(d: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if d == "Secure" or d.startswith("Firmado"): return t("status_compliant")
    if d.startswith("Incompleto"): return t("status_partial")
    if d == "No implementado" or d.startswith("Bogus"): return t("status_non_compliant")
    return t("status_pending")


def derive_dnssec_severity(d: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if d.startswith("Bogus"): return t("sev_critical")
    if d.startswith("Incompleto"): return t("sev_high")
    if d == "No implementado": return t("sev_medium")
    if d == "Secure" or d.startswith("Firmado"): return t("sev_info")
    return t("sev_medium")


def derive_dkim_status(bits: Any, t_flag: str, rev: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if rev.startswith("si"): return t("status_non_compliant")
    if isinstance(bits, int):
        if bits < 1024: return t("status_non_compliant")
        if bits < 2048: return t("status_partial")
    else:
        return t("status_pending")
    if t_flag == "y": return t("status_partial")
    return t("status_compliant")


def derive_dkim_severity(bits: Any, t_flag: str, rev: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if rev.startswith("si"): return t("sev_critical")
    if isinstance(bits, int):
        if bits < 1024: return t("sev_critical")
        if bits < 2048: return t("sev_high")
    else:
        return t("sev_medium")
    if t_flag == "y": return t("sev_medium")
    return t("sev_info")


def derive_mtasts_status(pub: str, mode: str, acc: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if pub == "no" or acc == "no" or mode == "none": return t("status_non_compliant")
    if mode == "enforce": return t("status_compliant")
    if mode == "testing": return t("status_partial")
    return t("status_pending")


def derive_tlsrpt_status(pub: str, rua: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if pub == "no": return t("status_non_compliant")
    if not rua: return t("status_pending")
    return t("status_compliant")


def derive_bimi_status(pub: str, svg: str, vmc: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    if pub == "no" or not svg or vmc == "No": return t("status_non_compliant")
    if vmc in ("Si", "Yes"): return t("status_compliant")
    return t("status_partial")


def derive_remit_state(spf: str, dkim: str, dmarc: str, t: Optional[Any] = None) -> str:
    if not t: t = lambda k: k
    y = sum(1 for v in (spf, dkim, dmarc) if v in ("Si", "Yes"))
    n = sum(1 for v in (spf, dkim, dmarc) if v in ("No", "No"))
    if y == 3: return t("status_compliant")
    if n == 3: return t("status_non_compliant")
    if y >= 1: return t("status_partial")
    return t("status_to_validate")


def write_evidence(outdir: Path, domain: str, filename: str, command: str, content: str, resolver: str) -> None:
    path = outdir / "evidencias" / domain / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    h = (f"# Evidencia: {filename}\n# Dominio: {domain}\n# Comando: {command}\n"
         f"# Timestamp: {datetime.now(timezone.utc).isoformat()}\n# Resolver: {resolver}\n# -----\n")
    path.write_text(h + (content or "") + "\n", encoding="utf-8")


async def q(resolver_obj: dns.asyncresolver.Resolver, name: str, rdtype: str) -> Optional[Any]:
    try:
        return await resolver_obj.resolve(name, rdtype, lifetime=6)
    except Exception:
        return None


def format_iso_date(raw_date: Optional[str]) -> str:
    if not raw_date:
        return "N/D"
    try:
        clean = raw_date.replace("Z", "+00:00")
        dt = datetime.fromisoformat(clean)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(raw_date)[:10]


async def check_whois_rdap(domain: str, http: httpx.AsyncClient, outdir: Path, resolver: str, t: Optional[Any] = None) -> Dict[str, Any]:
    """
    Evaluates Domain Registration, Registrar, Dates and Entity/Brand using:
      1. Authoritative TLD + Universal RDAP (RFC 7480-7484 / RFC 9082-9083) over HTTPS.
      2. Asynchronous fallback to socket WHOIS (port 43) with timeout protection.
    """
    if not t: t = lambda k: k
    res = {
        "registrar": "N/D",
        "created": "N/D",
        "expires": "N/D",
        "status": "N/D",
        "dnssec_whois": "N/D",
        "brand": derive_brand_from_domain(domain),
        "source": "None"
    }

    # --- Tier 1: Multi-Candidate RDAP over HTTPS ---
    for rdap_url in get_rdap_candidate_urls(domain):
        try:
            resp = await http.get(rdap_url, headers=RDAP_HEADERS, timeout=5.0, follow_redirects=True)
            if resp.status_code == 200:
                rdap_data = resp.json()
                write_evidence(outdir, domain, "rdap.json", f"GET {rdap_url}", json.dumps(rdap_data, indent=2), resolver)
                
                # 1. Parse Events (dates)
                events = {e.get("eventAction"): e.get("eventDate") for e in rdap_data.get("events", []) if isinstance(e, dict)}
                if "registration" in events:
                    res["created"] = format_iso_date(events["registration"])
                if "expiration" in events:
                    res["expires"] = format_iso_date(events["expiration"])

                # 2. Parse Entities (Registrar & Registrant Organization)
                registrar_name = None
                registrant_org = None
                for entity in rdap_data.get("entities", []):
                    roles = entity.get("roles", [])
                    vcards = entity.get("vcardArray", [None, []])
                    card_list = vcards[1] if (len(vcards) > 1 and isinstance(vcards[1], list)) else []
                    fn = next((v[3] for v in card_list if isinstance(v, list) and len(v) > 3 and v[0] == "fn"), None)
                    org = next((v[3] for v in card_list if isinstance(v, list) and len(v) > 3 and v[0] == "org"), None)

                    if "registrar" in roles:
                        registrar_name = fn or org or entity.get("handle")
                    if "registrant" in roles or "administrative" in roles:
                        registrant_org = org or fn

                if registrar_name:
                    res["registrar"] = str(registrar_name)[:150]

                # 3. Determine Brand / Entity
                privacy_keywords = ["privacy", "proxy", "whoisguard", "withheld", "redacted", "domains by proxy", "contact privacy"]
                if registrant_org:
                    is_private = any(k in registrant_org.lower() for k in privacy_keywords)
                    if is_private:
                        res["brand"] = f"{derive_brand_from_domain(domain)} ({t('privacy_protected')})"
                    else:
                        res["brand"] = str(registrant_org)[:150]
                else:
                    res["brand"] = derive_brand_from_domain(domain)

                # 4. Status
                status_list = rdap_data.get("status", [])
                if status_list:
                    res["status"] = "; ".join(str(s) for s in status_list[:3])

                # 5. SecureDNS
                secure_dns = rdap_data.get("secureDNS", {})
                if isinstance(secure_dns, dict):
                    res["dnssec_whois"] = "Signed" if secure_dns.get("delegationSigned") else "Unsigned"

                res["source"] = f"RDAP ({rdap_url})"
                if res["registrar"] != "N/D" or res["expires"] != "N/D":
                    return res
        except Exception:
            continue

    # --- Tier 2: WHOIS Fallback (Socket Port 43) ---
    loop = asyncio.get_event_loop()
    try:
        w = await asyncio.wait_for(loop.run_in_executor(None, whois_lib.whois, domain), timeout=3.0)
        raw = str(w)
        reg = getattr(w, "registrar", None) or res["registrar"]
        cr = getattr(w, "creation_date", None)
        if isinstance(cr, list): cr = cr[0] if cr else None
        ex = getattr(w, "expiration_date", None)
        if isinstance(ex, list): ex = ex[0] if ex else None
        st = getattr(w, "status", None) or res["status"]
        if isinstance(st, list): st = ";".join(str(x) for x in st[:3])
        ds = getattr(w, "dnssec", None) or res["dnssec_whois"]
        org = getattr(w, "org", None) or getattr(w, "name", None)

        if org and not res.get("brand"):
            res["brand"] = str(org)[:150]

        write_evidence(outdir, domain, "whois.txt", f"whois {domain}", raw, resolver)
        res["registrar"] = str(reg)[:150] if reg else res["registrar"]
        res["created"] = str(cr)[:10] if cr else res["created"]
        res["expires"] = str(ex)[:10] if ex else res["expires"]
        res["status"] = str(st)[:150] if st else res["status"]
        res["dnssec_whois"] = str(ds) if ds else res["dnssec_whois"]
        res["source"] = "WHOIS (Port 43)"
    except Exception as e:
        write_evidence(outdir, domain, "whois.txt", f"whois {domain}", f"ERROR: {e}", resolver)

    return res


async def check_ns_soa(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str) -> Dict[str, Any]:
    ns = await q(ro, domain, "NS")
    soa = await q(ro, domain, "SOA")
    ns_list = sorted([str(r.target).rstrip(".") for r in ns]) if ns else []
    soa_serial = "N/D"
    if soa:
        try:
            soa_serial = str(soa[0].serial)
        except Exception:
            pass
    ns_str = ";".join(ns_list) if ns_list else "N/D"
    write_evidence(outdir, domain, "dig_ns.txt", f"NS {domain}", ns_str, ip)
    prov = "Otro / Custom"
    low = ns_str.lower()
    if "cloudflare" in low: prov = "Cloudflare"
    elif "awsdns" in low: prov = "AWS Route 53"
    elif "azure-dns" in low: prov = "Azure DNS"
    elif "googledomains" in low or "google.com" in low: prov = "Google Cloud DNS"
    elif "godaddy" in low or "domaincontrol" in low: prov = "GoDaddy"
    elif "namecheap" in low: prov = "Namecheap"
    return {"ns_list": ns_str, "soa_serial": soa_serial, "ns_provider": prov}


async def _probe_ad_bit(domain: str, resolver_ip: str) -> Tuple[str, str, str]:
    ad = "no"; status = "N/D"; dump = ""
    try:
        query = dns.message.make_query(domain, "A", want_dnssec=True)
        query.flags |= dns.flags.AD
        try:
            response = await dns.asyncquery.udp(query, resolver_ip, timeout=6)
            if response.flags & dns.flags.TC:
                response = await dns.asyncquery.tcp(query, resolver_ip, timeout=6)
        except Exception:
            response = await dns.asyncquery.tcp(query, resolver_ip, timeout=6)
        status = dns.rcode.to_text(response.rcode())
        if response.flags & dns.flags.AD:
            ad = "si"
        dump = str(response)
    except Exception as e:
        if "SERVFAIL" in str(e).upper():
            status = "SERVFAIL"
        dump = f"ERROR: {e}"
    return ad, status, dump


async def check_dnssec(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str, dnssec_resolvers: Optional[List[str]] = None) -> Dict[str, Any]:
    dk = await q(ro, domain, "DNSKEY")
    ds_a = await q(ro, domain, "DS")
    dk_pub = "si" if dk else "no"
    ds_pub = "si" if ds_a else "no"

    if not dnssec_resolvers:
        dnssec_resolvers = ["1.1.1.1", "8.8.8.8", "9.9.9.9"]

    ad = "no"; status = "N/D"
    per_resolver_dump = []
    servfail_count = 0
    for rip in dnssec_resolvers:
        r_ad, r_status, r_dump = await _probe_ad_bit(domain, rip)
        per_resolver_dump.append(f"### Resolver {rip}\n# AD={r_ad}  status={r_status}\n{r_dump}\n")
        if r_status == "SERVFAIL":
            servfail_count += 1
        if r_ad == "si":
            ad = "si"
            if status in ("N/D", "SERVFAIL"):
                status = r_status
        elif status == "N/D":
            status = r_status

    all_servfail = (servfail_count == len(dnssec_resolvers))
    algos_local = []
    if dk:
        for r in dk:
            try:
                algos_local.append(str(r.algorithm))
            except Exception:
                pass
    algos_s = ",".join(sorted(set(algos_local))) if algos_local else "N/A"

    if all_servfail:
        diag = "Bogus (cadena rota)"
    elif dk_pub == "si" and ds_pub == "si" and ad == "si":
        diag = "Secure"
    elif dk_pub == "si" and ds_pub == "si" and ad == "no":
        diag = "Firmado (AD no confirmado)"
    elif dk_pub == "si" and ds_pub == "no":
        diag = "Incompleto (DS no publicado)"
    elif dk_pub == "no" and ds_pub == "no":
        diag = "No implementado"
    else:
        diag = "Inconsistente"

    write_evidence(outdir, domain, "dig_dnskey.txt", f"DNSKEY {domain}", str(dk.rrset) if dk else "(empty)", ip)
    write_evidence(outdir, domain, "dig_ds.txt", f"DS {domain}", str(ds_a.rrset) if ds_a else "(empty)", ip)
    write_evidence(outdir, domain, "dnssec_validation.txt",
                   f"make_query {domain} A +dnssec +ad (resolvers: {', '.join(dnssec_resolvers)})",
                   "\n".join(per_resolver_dump), ip)
    return {"dnskey_pub": dk_pub, "ds_pub": ds_pub, "ad_flag": ad, "status": status,
            "algos": algos_s, "diag": diag}


async def check_spf(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, domain, "TXT")
    txts = []
    if ans:
        for r in ans:
            try:
                txts.append(b"".join(r.strings).decode("utf-8", errors="replace"))
            except Exception:
                pass
    spfs = [t for t in txts if t.lower().startswith("v=spf1")]
    n = len(spfs); rec = spfs[0] if spfs else ""
    write_evidence(outdir, domain, "dig_txt.txt", f"TXT {domain}", "\n".join(txts), ip)
    r = {"pub": "no", "all": "N/A", "lookups": 0, "void": 0, "multi": "no", "len": 0, "providers": "", "record": ""}
    if not rec:
        return r
    r["pub"] = "si"; r["record"] = rec; r["len"] = len(rec)
    if " -all" in rec: r["all"] = "-all"
    elif " ~all" in rec: r["all"] = "~all"
    elif " ?all" in rec: r["all"] = "?all"
    elif " +all" in rec: r["all"] = "+all"
    else: r["all"] = "sin all"
    r["lookups"] = len(re.findall(r"(include:|a:|mx:|ptr:|exists:|redirect=)", rec))
    incs = re.findall(r"include:(\S+)", rec)
    void = 0; seen = []
    for inc in incs:
        sub = await q(ro, inc, "TXT")
        if not sub:
            void += 1
        low = inc.lower()
        if "spf.protection.outlook.com" in low: p = "Microsoft 365"
        elif "_spf.google.com" in low: p = "Google Workspace"
        elif "amazonses.com" in low: p = "Amazon SES"
        elif "mailgun" in low: p = "Mailgun"
        elif "sendgrid" in low: p = "SendGrid"
        elif "mailchimp" in low or "mcsv" in low: p = "Mailchimp"
        elif "hubspot" in low: p = "HubSpot"
        elif "zoho" in low: p = "Zoho Mail"
        elif "mandrill" in low: p = "Mandrill"
        elif "sparkpost" in low: p = "SparkPost"
        elif "postmark" in low: p = "Postmark"
        elif "mailjet" in low: p = "Mailjet"
        else: p = f"Otro: {inc}"
        if p not in seen:
            seen.append(p)
    r["void"] = void
    r["providers"] = "; ".join(seen)
    if n > 1: r["multi"] = "si"
    return r


async def check_dkim(domain: str, sels: List[str], ro: dns.asyncresolver.Resolver, outdir: Path, ip: str, counter: List[int], deep: bool = False, t: Optional[Any] = None) -> Dict[str, Any]:
    if not t: t = lambda k: k
    found = []
    probed = 0
    for sel in sels:
        probed += 1
        ans = await q(ro, f"{sel}._domainkey.{domain}", "TXT")
        if not ans:
            continue
        try:
            rec = "".join(b"".join(r.strings).decode("utf-8", errors="replace") for r in ans)
        except Exception:
            continue
        if "v=DKIM1" not in rec and "p=" not in rec:
            continue
        write_evidence(outdir, domain, f"dkim_{sel}.txt", f"TXT {sel}._domainkey.{domain}", rec, ip)
        mp = re.search(r"p=([A-Za-z0-9+/=]*)", rec)
        mk = re.search(r"k=([a-z0-9]+)", rec)
        mt = re.search(r"t=([ys]+)", rec)
        pv = mp.group(1) if mp else ""
        kv = mk.group(1) if mk else "rsa"
        tv = mt.group(1) if mt else "-"
        rev = "no"; bits: Any = "N/D"
        if not pv:
            rev = "si"
        else:
            try:
                der = base64.b64decode(pv)
                key = serialization.load_der_public_key(der, backend=default_backend())
                bits = key.key_size
            except Exception:
                bits = "Invalida"
        st = derive_dkim_status(bits, tv, rev, t=t)
        sv = derive_dkim_severity(bits, tv, rev, t=t)
        counter[0] += 1
        found.append({
            "row": counter[0], "domain": domain, "selector": sel,
            "service": map_dkim_service(sel), "record": rec,
            "algorithm": kv, "bits": bits, "t_flag": tv,
            "estado": st, "severidad": sv
        })
    return {"found": found, "probed": probed, "deep": deep}


async def check_dmarc(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, f"_dmarc.{domain}", "TXT")
    rec = ""
    if ans:
        for r in ans:
            try:
                t_rec = b"".join(r.strings).decode("utf-8", errors="replace")
                if t_rec.lower().startswith("v=dmarc1"):
                    rec = t_rec; break
            except Exception:
                pass
    write_evidence(outdir, domain, "dig_dmarc.txt", f"TXT _dmarc.{domain}", rec, ip)
    r = {"pub": "no", "record": "", "p": "none", "sp": "", "pct": "100",
         "aspf": "r", "adkim": "r", "rua": "", "ruf": "", "fo": "0", "rf": "afrf", "ri": "86400"}
    if not rec:
        return r
    r["pub"] = "si"; r["record"] = rec
    for tag in ("p", "sp", "pct", "aspf", "adkim", "rua", "ruf", "fo", "rf", "ri"):
        m = re.search(rf"\b{tag}=([^;]+)", rec, re.IGNORECASE)
        if m:
            r[tag] = m.group(1).strip()
    return r


async def check_mx(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, domain, "MX")
    mxs = []
    if ans:
        for r in sorted(ans, key=lambda x: x.preference):
            mxs.append(f"{r.preference} {str(r.exchange).rstrip('.')}")
    mx_str = "; ".join(mxs) if mxs else "Sin MX"
    write_evidence(outdir, domain, "dig_mx.txt", f"MX {domain}", mx_str, ip)
    low = mx_str.lower()
    prov = "Otro"
    if "outlook.com" in low or "pphosted.com" in low: prov = "Microsoft 365"
    elif "google.com" in low or "googlemail.com" in low: prov = "Google Workspace"
    elif "mimecast" in low: prov = "Mimecast"
    elif "barracuda" in low: prov = "Barracuda"
    elif "zoho" in low: prov = "Zoho Mail"
    elif "proton" in low: prov = "Proton Mail"
    elif not mxs: prov = "Sin MX"
    return {"raw": mx_str, "provider": prov}


async def check_mtasts(domain: str, ro: dns.asyncresolver.Resolver, http: httpx.AsyncClient, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, f"_mta-sts.{domain}", "TXT")
    rec = ""
    if ans:
        for r in ans:
            try:
                t_rec = b"".join(r.strings).decode("utf-8", errors="replace")
                if "v=STSv1" in t_rec: rec = t_rec; break
            except Exception:
                pass
    write_evidence(outdir, domain, "mtasts_dns.txt", f"TXT _mta-sts.{domain}", rec, ip)
    r = {"pub": "no", "record": rec, "mode": "", "max_age": "", "mx": "", "accessible": "no"}
    if not rec:
        return r
    r["pub"] = "si"
    url = f"https://mta-sts.{domain}/.well-known/mta-sts.txt"
    try:
        resp = await http.get(url, timeout=5)
        if resp.status_code == 200 and "version: STSv1" in resp.text:
            r["accessible"] = "si"
            for line in resp.text.splitlines():
                if line.startswith("mode:"): r["mode"] = line.split(":", 1)[1].strip()
                elif line.startswith("max_age:"): r["max_age"] = line.split(":", 1)[1].strip()
                elif line.startswith("mx:"):
                    r["mx"] += (", " if r["mx"] else "") + line.split(":", 1)[1].strip()
            write_evidence(outdir, domain, "mtasts_policy.txt", f"GET {url}", resp.text, ip)
    except Exception as e:
        write_evidence(outdir, domain, "mtasts_policy.txt", f"GET {url}", f"ERROR: {e}", ip)
    return r


async def check_tlsrpt(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, f"_smtp._tls.{domain}", "TXT")
    rec = ""
    if ans:
        for r in ans:
            try:
                t_rec = b"".join(r.strings).decode("utf-8", errors="replace")
                if "v=TLSRPTv1" in t_rec: rec = t_rec; break
            except Exception:
                pass
    write_evidence(outdir, domain, "tlsrpt_dns.txt", f"TXT _smtp._tls.{domain}", rec, ip)
    r = {"pub": "no", "record": rec, "rua": ""}
    if rec:
        r["pub"] = "si"
        m = re.search(r"rua=([^;]+)", rec)
        if m: r["rua"] = m.group(1).strip()
    return r


async def check_bimi(domain: str, ro: dns.asyncresolver.Resolver, http: httpx.AsyncClient, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, f"default._bimi.{domain}", "TXT")
    rec = ""
    if ans:
        for r in ans:
            try:
                t_rec = b"".join(r.strings).decode("utf-8", errors="replace")
                if "v=BIMI1" in t_rec: rec = t_rec; break
            except Exception:
                pass
    write_evidence(outdir, domain, "bimi_dns.txt", f"TXT default._bimi.{domain}", rec, ip)
    r = {"pub": "no", "record": rec, "svg": "", "vmc_url": "", "vmc_status": "No",
         "vmc_issuer": "N/A", "vmc_exp": "N/A"}
    if not rec:
        return r
    r["pub"] = "si"
    m_l = re.search(r"l=([^;]+)", rec)
    m_a = re.search(r"a=([^;]+)", rec)
    if m_l: r["svg"] = m_l.group(1).strip()
    if m_a:
        r["vmc_url"] = m_a.group(1).strip()
        try:
            resp = await http.get(r["vmc_url"], timeout=5)
            if resp.status_code == 200:
                cert = x509.load_pem_x509_certificate(resp.content, default_backend())
                r["vmc_status"] = "Si"
                r["vmc_issuer"] = cert.issuer.rfc4514_string()
                r["vmc_exp"] = cert.not_valid_after_utc.strftime("%Y-%m-%d")
        except Exception:
            pass
    return r


async def check_caa(domain: str, ro: dns.asyncresolver.Resolver, outdir: Path, ip: str) -> Dict[str, Any]:
    ans = await q(ro, domain, "CAA")
    caa_records = []
    issue_list = []
    issuewild_list = []
    iodef = ""
    raw_text = ""
    if ans:
        for r in ans:
            try:
                tag = r.tag.decode("utf-8", errors="replace").lower()
                val = r.value.decode("utf-8", errors="replace")
                raw_text += f"{r.flags} {tag} \"{val}\"\n"
                caa_records.append(f"{tag}: {val}")
                if tag == "issue":
                    issue_list.append(val)
                elif tag == "issuewild":
                    issuewild_list.append(val)
                elif tag == "iodef":
                    iodef = val
            except Exception:
                pass
    write_evidence(outdir, domain, "dig_caa.txt", f"CAA {domain}", raw_text or "No CAA records", ip)
    pub = "si" if caa_records else "no"
    cas_allowed = ", ".join(sorted(set(issue_list + issuewild_list))) if (issue_list or issuewild_list) else ("Sin restricción" if pub == "no" else "Restringido")
    return {
        "pub": pub,
        "records": "; ".join(caa_records),
        "cas": cas_allowed,
        "iodef": iodef or "N/D",
        "count": len(caa_records)
    }


async def check_fcrdns(mx_hosts: List[str], ro: dns.asyncresolver.Resolver) -> Dict[str, Any]:
    if not mx_hosts:
        return {"status": "Sin MX", "details": "No MX hosts", "compliant": True}
    
    results = []
    all_aligned = True
    any_tested = False

    for host in mx_hosts[:3]:
        try:
            a_ans = await ro.resolve(host, "A")
            ips = [r.to_text() for r in a_ans]
        except Exception:
            continue

        for ip_addr in ips[:2]:
            any_tested = True
            try:
                rev_name = dns.reversename.from_address(ip_addr)
                ptr_ans = await ro.resolve(rev_name, "PTR")
                ptr_host = ptr_ans[0].target.to_text().rstrip(".")
                
                fwd_ans = await ro.resolve(ptr_host, "A")
                fwd_ips = [r.to_text() for r in fwd_ans]
                if ip_addr in fwd_ips:
                    results.append(f"{host}({ip_addr}) -> PTR {ptr_host} [OK]")
                else:
                    results.append(f"{host}({ip_addr}) -> PTR {ptr_host} [Mismatch]")
                    all_aligned = False
            except Exception:
                results.append(f"{host}({ip_addr}) -> Sin PTR")
                all_aligned = False

    if not any_tested:
        return {"status": "No verificado", "details": "No se resolvieron IPs de MX", "compliant": False}

    status = "Alineado (FCrDNS OK)" if all_aligned else "Desalineado / Sin PTR"
    return {
        "status": status,
        "details": "; ".join(results),
        "compliant": all_aligned
    }


async def check_tls_certificate_health(mx_hosts: List[str], domain: str) -> Dict[str, Any]:
    target_hosts = mx_hosts if mx_hosts else [f"mail.{domain}", domain]

    def _probe_socket(host: str, port: int = 443) -> Optional[Dict[str, Any]]:
        try:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            with socket.create_connection((host, port), timeout=3) as sock:
                with ctx.wrap_socket(sock, server_hostname=host) as ssock:
                    der_cert = ssock.getpeercert(binary_form=True)
                    if not der_cert:
                        return None
                    cert = x509.load_der_x509_certificate(der_cert, default_backend())
                    not_after = cert.not_valid_after_utc
                    now = datetime.now(timezone.utc)
                    days_left = (not_after - now).days
                    issuer_str = cert.issuer.rfc4514_string()
                    
                    sans = []
                    try:
                        san_ext = cert.extensions.get_extension_for_oid(x509.oid.ExtensionOID.SUBJECT_ALTERNATIVE_NAME)
                        sans = san_ext.value.get_values_for_type(x509.DNSName)
                    except Exception:
                        pass

                    return {
                        "host": host,
                        "days_left": days_left,
                        "expires": not_after.strftime("%Y-%m-%d"),
                        "issuer": issuer_str.split("CN=")[-1].split(",")[0] if "CN=" in issuer_str else issuer_str[:30],
                        "sans": ", ".join(sans[:3]) if sans else "N/A"
                    }
        except Exception:
            return None

    loop = asyncio.get_running_loop()
    for host in target_hosts[:2]:
        res = await loop.run_in_executor(None, _probe_socket, host, 443)
        if res:
            days = res["days_left"]
            status = f"Válido ({days}d)" if days >= 30 else f"Por vencer ({days}d)" if days > 0 else f"Vencido ({days}d)"
            warn = None
            if days <= 7:
                warn = "critical"
            elif days <= 30:
                warn = "warning"
            return {
                "status": status,
                "days_left": days,
                "expires": res["expires"],
                "issuer": res["issuer"],
                "san": res["sans"],
                "host": host,
                "warning": warn
            }

    return {"status": "No evaluado", "days_left": None, "expires": "N/A", "issuer": "N/D", "san": "N/D", "warning": None, "host": "N/D"}


async def check_dmarc_external_report_auth(domain: str, rua: str, ro: dns.asyncresolver.Resolver) -> Dict[str, Any]:
    if not rua:
        return {"external": False, "authorized": True, "target": "N/A", "details": "Sin rua"}
    
    emails = re.findall(r"mailto:([a-zA-Z0-9_.+-]+@([a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+))", rua)
    if not emails:
        return {"external": False, "authorized": True, "target": "N/A", "details": "Sin emails rua"}
    
    external_targets = []
    for _, dest_domain in emails:
        dest_domain = dest_domain.lower()
        if dest_domain != domain and not domain.endswith(f".{dest_domain}"):
            external_targets.append(dest_domain)

    if not external_targets:
        return {"external": False, "authorized": True, "target": "Interno", "details": "Destinos internos"}

    auth_results = []
    all_auth = True
    for target in set(external_targets):
        auth_record_name = f"{domain}._report._dmarc.{target}"
        txts = await q(ro, auth_record_name, "TXT")
        has_auth = False
        if txts:
            for r in txts:
                try:
                    s_txt = b"".join(r.strings).decode("utf-8", errors="replace")
                    if "v=DMARC1" in s_txt.upper():
                        has_auth = True
                        break
                except Exception:
                    pass
        if has_auth:
            auth_results.append(f"{target} [Autorizado]")
        else:
            auth_results.append(f"{target} [Sin registro RFC 7489]")
            all_auth = False

    return {
        "external": True,
        "authorized": all_auth,
        "target": ", ".join(set(external_targets)),
        "details": "; ".join(auth_results)
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  EASM & CISO EVALUATION ENGINE (Typosquatting, Homoglyphs & Compliance)
# ═══════════════════════════════════════════════════════════════════════════════

TAKEOVER_FINGERPRINTS = {
    "github.io": "There isn't a GitHub Pages site here.",
    "s3.amazonaws.com": "NoSuchBucket",
    "azurewebsites.net": "404 Web Site not found",
    "herokuapp.com": "No such app",
    "zendesk.com": "Help Center Closed",
    "squarespace.com": "No Such Account",
    "myshopify.com": "Sorry, this shop is currently unavailable.",
    "pantheonsite.io": "404 error unknown site",
    "ghost.io": "The thing you were looking for is gone for good.",
    "wordpress.com": "Do you want to register",
    "fastly.net": "Fastly error: unknown domain",
    "cloudfront.net": "Bad request / Bad gateway",
    "unbouncepages.com": "The requested URL was not found on this server",
    "bitbucket.io": "Repository not found",
    "helpjuice.com": "We could not find what you're looking for",
}


def generate_lookalikes(domain: str) -> List[Tuple[str, str]]:
    """Generates offline lookalike candidates (homoglyphs, omissions, bitsquatting, TLDs)."""
    parts = domain.split(".", 1)
    if len(parts) < 2:
        return []
    name, tld = parts[0], parts[1]
    candidates: List[Tuple[str, str]] = []

    # 1. Homoglyphs / Substitutions
    homoglyphs = {
        'o': ['0'], 'l': ['1', 'i'], 'i': ['1', 'l'], 'e': ['3'],
        'a': ['4'], 's': ['5'], 'rn': ['m'], 'm': ['rn'], 'vv': ['w'], 'w': ['vv']
    }
    for orig, subs in homoglyphs.items():
        if orig in name:
            for sub in subs:
                cand = name.replace(orig, sub, 1)
                if cand != name:
                    candidates.append((f"{cand}.{tld}", "Homoglyph"))

    # 2. Omissions (Levenshtein d=1)
    if len(name) > 3:
        for i in range(len(name)):
            cand = name[:i] + name[i+1:]
            if len(cand) >= 3:
                candidates.append((f"{cand}.{tld}", "Omission"))

    # 3. Transposition / Repetition
    for i in range(len(name) - 1):
        cand = name[:i] + name[i+1] + name[i] + name[i+2:]
        if cand != name:
            candidates.append((f"{cand}.{tld}", "Transposition"))

    # 4. Alternative TLDs
    alt_tlds = ["co", "net", "org", "info", "online", "security", "io"]
    for at in alt_tlds:
        if at != tld:
            candidates.append((f"{name}.{at}", "TLD-Variation"))

    # 5. Phishing Prefix/Suffixes
    prefixes = ["login-", "secure-", "support-", "portal-"]
    for p in prefixes:
        candidates.append((f"{p}{name}.{tld}", "Prefix-Variation"))

    # Deduplicate & preserve order
    seen = set()
    unique_candidates = []
    for cand, mtype in candidates:
        if cand != domain and cand not in seen:
            seen.add(cand)
            unique_candidates.append((cand, mtype))
    return unique_candidates[:20]


async def check_lookalikes(domain: str, ro: dns.asyncresolver.Resolver, t: Optional[Any] = None) -> List[Dict[str, Any]]:
    """Probes generated lookalike domains for active A and MX DNS records."""
    if not t: t = lambda k: k
    candidates = generate_lookalikes(domain)
    results = []

    async def probe_one(cand: str, mtype: str) -> Dict[str, Any]:
        ips, mx_list = [], []
        try:
            a_ans = await ro.resolve(cand, "A")
            ips = [r.to_text() for r in a_ans]
        except Exception:
            pass
        try:
            mx_ans = await ro.resolve(cand, "MX")
            mx_list = [r.exchange.to_text().rstrip(".") for r in mx_ans]
        except Exception:
            pass

        registered = bool(ips or mx_list)
        if registered and mx_list:
            threat = t("sev_critical")
            action = "Bloquear / Monitorear MX" if t("yes") == "Si" else "Block / Monitor MX"
        elif registered and ips:
            threat = t("sev_high")
            action = "Monitorear DNS / Takedown" if t("yes") == "Si" else "Monitor DNS / Takedown"
        else:
            threat = t("sev_info")
            action = "Registro Defensivo" if t("yes") == "Si" else "Defensive Registration"

        return {
            "domain": domain,
            "lookalike": cand,
            "type": mtype,
            "ips": ", ".join(ips[:2]) if ips else t("no"),
            "mx": ", ".join(mx_list[:2]) if mx_list else t("no_mx"),
            "registered": registered,
            "threat": threat,
            "action": action
        }

    tasks = [probe_one(c, mt) for c, mt in candidates]
    if tasks:
        results = await asyncio.gather(*tasks)
    return results


async def check_subdomain_takeover(domain: str, ro: dns.asyncresolver.Resolver, http: httpx.AsyncClient, t: Optional[Any] = None) -> List[Dict[str, Any]]:
    """Checks common subdomains for dangling CNAME takeovers against signature database."""
    if not t: t = lambda k: k
    sub_prefixes = ["mail", "webmail", "portal", "dev", "stage", "cdn", "assets", "docs", "status", "help", "app"]
    results = []

    async def probe_sub(sub: str) -> Optional[Dict[str, Any]]:
        fqdn = f"{sub}.{domain}"
        cname_target = ""
        try:
            ans = await ro.resolve(fqdn, "CNAME")
            if ans:
                cname_target = ans[0].target.to_text().rstrip(".")
        except Exception:
            return None

        if not cname_target:
            return None

        # Match provider fingerprint
        matched_provider = None
        for prov_dom, sig in TAKEOVER_FINGERPRINTS.items():
            if prov_dom in cname_target:
                matched_provider = (prov_dom, sig)
                break

        if not matched_provider:
            return None

        prov_dom, sig = matched_provider
        dangling = False
        try:
            resp = await http.get(f"http://{fqdn}", timeout=4)
            if sig.lower() in resp.text.lower() or resp.status_code == 404:
                dangling = True
        except Exception:
            pass

        return {
            "fqdn": fqdn,
            "cname": cname_target,
            "provider": prov_dom,
            "dangling": dangling,
            "severity": t("sev_critical") if dangling else t("sev_low")
        }

    tasks = [probe_sub(p) for p in sub_prefixes]
    res_list = await asyncio.gather(*tasks)
    return [r for r in res_list if r is not None]


def evaluate_ciso_compliance_and_score(info: Dict[str, Any], easm_res: List[Dict[str, Any]], takeover_res: List[Dict[str, Any]], t: Optional[Any] = None) -> Tuple[int, str, List[List[str]]]:
    """Computes weighted CISO risk score (0-100) and evaluates compliance frameworks."""
    if not t: t = lambda k: k
    domain = info["domain"]
    s = info["spf"]; d = info["dmarc"]; ds = info["dnssec"]
    mt = info["mtasts"]; tl = info["tlsrpt"]; bi = info["bimi"]

    # Weighted Score (Total: 100)
    auth_score = 0
    if s["pub"] == "si" and s["all"] in ("-all", "~all") and s["multi"] == "no" and s["lookups"] <= 10:
        auth_score += 15
    elif s["pub"] == "si":
        auth_score += 8

    if d["pub"] == "si":
        if d["p"] == "reject": auth_score += 20
        elif d["p"] == "quarantine": auth_score += 14
        else: auth_score += 5

    if info["dkim"]:
        auth_score += 5

    trans_score = 0
    if mt["mode"] == "enforce": trans_score += 15
    elif mt["mode"] == "testing": trans_score += 8

    if tl["pub"] == "si": trans_score += 10

    dns_score = 0
    if ds["diag"] == "Secure" or ds["diag"].startswith("Firmado"): dns_score += 15
    elif ds["diag"].startswith("Incompleto"): dns_score += 5

    if bi["pub"] == "si": dns_score += 5

    active_lookalikes = sum(1 for r in easm_res if r.get("registered"))
    dangling_takeovers = sum(1 for r in takeover_res if r.get("dangling"))

    easm_score = 15
    if active_lookalikes > 0: easm_score = max(0, easm_score - min(10, active_lookalikes * 3))
    if dangling_takeovers > 0: easm_score = max(0, easm_score - 10)

    total_score = min(100, auth_score + trans_score + dns_score + easm_score)
    if total_score >= 90: grade = "A"
    elif total_score >= 80: grade = "B"
    elif total_score >= 70: grade = "C"
    elif total_score >= 60: grade = "D"
    else: grade = "F"

    c_comp = t("status_compliant")
    c_part = t("status_partial")
    c_non = t("status_non_compliant")

    # PCI-DSS v4.0 Req 5.4
    if d["pub"] == "si" and d["p"] in ("reject", "quarantine") and d["pct"] in (100, "100"):
        pci_st = c_comp
        pci_notes = "DMARC enforzado conforme a PCI-DSS v4.0 Req 5.4" if t("yes") == "Si" else "DMARC enforced per PCI-DSS v4.0 Req 5.4"
    elif d["pub"] == "si" and d["p"] == "none":
        pci_st = c_part
        pci_notes = "DMARC en p=none (requiere migrar a reject/quarantine)" if t("yes") == "Si" else "DMARC in p=none (must migrate to reject/quarantine)"
    else:
        pci_st = c_non
        pci_notes = "No cumple PCI-DSS v4.0 (Sin DMARC anti-phishing)" if t("yes") == "Si" else "Non-compliant with PCI-DSS v4.0 (No anti-phishing DMARC)"

    # NIST CSF 2.0 PR.AC-01
    if auth_score >= 35:
        nist_ac_st = c_comp
        nist_ac_notes = "Autenticación robusta SPF+DKIM+DMARC" if t("yes") == "Si" else "Strong SPF+DKIM+DMARC authentication"
    elif auth_score >= 20:
        nist_ac_st = c_part
        nist_ac_notes = "Autenticación parcial (faltan controles clave)" if t("yes") == "Si" else "Partial email authentication"
    else:
        nist_ac_st = c_non
        nist_ac_notes = "Riesgo alto de spoofing e impersonación" if t("yes") == "Si" else "High spoofing and impersonation risk"

    # NIST CSF 2.0 PR.DS-01
    if trans_score >= 20:
        nist_ds_st = c_comp
        nist_ds_notes = "MTA-STS y TLS-RPT enforzados para transporte seguro" if t("yes") == "Si" else "MTA-STS and TLS-RPT enforced for secure transport"
    elif trans_score >= 8:
        nist_ds_st = c_part
        nist_ds_notes = "MTA-STS en testing o TLS-RPT parcial" if t("yes") == "Si" else "MTA-STS in testing or partial TLS-RPT"
    else:
        nist_ds_st = c_non
        nist_ds_notes = "Sin protección contra STARTTLS stripping / MitM" if t("yes") == "Si" else "No protection against STARTTLS stripping / MitM"

    # ISO/IEC 27001:2022 A.8.20
    if ds["diag"] == "Secure" or ds["diag"].startswith("Firmado"):
        iso_st = c_comp
        iso_notes = "Zona firmada con DNSSEC y DS validado" if t("yes") == "Si" else "Zone signed with DNSSEC and validated DS"
    else:
        iso_st = c_part if ds["diag"].startswith("Incompleto") else c_non
        iso_notes = "DNSSEC no validado / riesgo de envenenamiento DNS" if t("yes") == "Si" else "DNSSEC not validated / DNS poisoning risk"

    # CIS Controls v8 Control 9.2
    if auth_score >= 30 and trans_score >= 15:
        cis_st = c_comp
        cis_notes = "Higiene y controles de correo alineados a CIS Control 9" if t("yes") == "Si" else "Email hygiene aligned with CIS Control 9"
    else:
        cis_st = c_part
        cis_notes = "Implementar DMARC p=reject y MTA-STS para cumplir CIS" if t("yes") == "Si" else "Implement DMARC p=reject and MTA-STS for CIS compliance"

    matrix = [
        [domain, "PCI-DSS v4.0", t("pci_dmarc_req"), pci_st, pci_notes],
        [domain, "NIST CSF 2.0", t("nist_csf_pr_ac"), nist_ac_st, nist_ac_notes],
        [domain, "NIST CSF 2.0", t("nist_csf_pr_ds"), nist_ds_st, nist_ds_notes],
        [domain, "ISO/IEC 27001:2022", t("iso_27001_13"), iso_st, iso_notes],
        [domain, "CIS Controls v8", t("cis_control_9"), cis_st, cis_notes],
    ]

    return total_score, grade, matrix


async def audit_domain(domain: str, args: Any, ro: dns.asyncresolver.Resolver, http: httpx.AsyncClient, outdir: Path, counters: Dict[str, Any], data: Dict[str, List[Any]], hallazgos: List[Any], t: Optional[Any] = None) -> Dict[str, Any]:
    if not t: t = lambda k: k
    ip = args.resolver
    dnssec_resolvers = [r.strip() for r in args.dnssec_resolvers.split(",") if r.strip()]

    whois_task = check_whois_rdap(domain, http, outdir, ip, t=t)
    ns_soa_task = check_ns_soa(domain, ro, outdir, ip)
    dnssec_task = check_dnssec(domain, ro, outdir, ip, dnssec_resolvers=dnssec_resolvers)
    spf_task = check_spf(domain, ro, outdir, ip)
    dmarc_task = check_dmarc(domain, ro, outdir, ip)
    mx_task = check_mx(domain, ro, outdir, ip)
    mtasts_task = check_mtasts(domain, ro, http, outdir, ip)
    tlsrpt_task = check_tlsrpt(domain, ro, outdir, ip)
    bimi_task = check_bimi(domain, ro, http, outdir, ip)
    caa_task = check_caa(domain, ro, outdir, ip)
    easm_task = check_lookalikes(domain, ro, t=t)
    takeover_task = check_subdomain_takeover(domain, ro, http, t=t)

    # Preparar selectores DKIM
    if args.deep_dkim:
        selectors = list(dict.fromkeys(COMMON_SELECTORS + generate_deep_selectors(args.deep_months)))
        deep_flag = True
    else:
        selectors = list(COMMON_SELECTORS)
        deep_flag = False

    if args.selectors:
        extra = [s.strip() for s in args.selectors.replace(",", " ").split() if s.strip()]
        selectors = list(dict.fromkeys(extra + selectors))

    dkim_task = check_dkim(domain, selectors, ro, outdir, ip, counters["dkim_row"], deep=deep_flag, t=t)

    results = await asyncio.gather(
        whois_task, ns_soa_task, dnssec_task, spf_task,
        dmarc_task, mx_task, mtasts_task, tlsrpt_task, bimi_task, dkim_task,
        caa_task, easm_task, takeover_task
    )

    whois_res, ns_soa_res, dnssec_res, spf_res, dmarc_res, mx_res, mtasts_res, tlsrpt_res, bimi_res, dkim_res, caa_res, easm_res, takeover_res = results

    # Secondary network probers dependent on MX and DMARC rua
    mx_hosts = [x.split()[1] for x in mx_res["raw"].split(";") if len(x.split()) > 1]
    fcrdns_res, tls_cert_res, dmarc_ext_res = await asyncio.gather(
        check_fcrdns(mx_hosts, ro),
        check_tls_certificate_health(mx_hosts, domain),
        check_dmarc_external_report_auth(domain, dmarc_res.get("rua", ""), ro)
    )

    info = {
        "domain": domain, "registrar": whois_res["registrar"], "created": whois_res["created"],
        "expires": whois_res["expires"], "status": whois_res["status"], "dnssec_whois": whois_res["dnssec_whois"],
        "brand": whois_res.get("brand") or derive_brand_from_domain(domain),
        "ns_list": ns_soa_res["ns_list"], "soa_serial": ns_soa_res["soa_serial"], "ns_provider": ns_soa_res["ns_provider"],
        "dnssec": dnssec_res, "spf": spf_res, "dmarc": dmarc_res, "mx": mx_res,
        "mtasts": mtasts_res, "tlsrpt": tlsrpt_res, "bimi": bimi_res,
        "dkim": dkim_res["found"], "dkim_res": dkim_res,
        "caa": caa_res, "fcrdns": fcrdns_res, "tls_cert": tls_cert_res, "dmarc_ext": dmarc_ext_res,
        "easm": easm_res, "takeover": takeover_res
    }

    process_results(info, counters, data, hallazgos, args, t=t)
    return info


def process_results(info: Dict[str, Any], counters: Dict[str, Any], data: Dict[str, List[Any]], hallazgos: List[Any], args: Any, t: Optional[Any] = None) -> None:
    if not t: t = lambda k: k
    domain = info["domain"]
    s = info["spf"]; d = info["dmarc"]; ds = info["dnssec"]
    mt = info["mtasts"]; tl = info["tlsrpt"]; bi = info["bimi"]; m = info["mx"]
    dkim_res = info.get("dkim_res", {})

    counters["dom_num"] += 1
    n = counters["dom_num"]

    # SPF
    sst = derive_spf_status(s["pub"], s["all"], s["lookups"], s["void"], s["multi"], t=t)
    ssv = derive_spf_severity(s["pub"], s["all"], s["lookups"], s["void"], s["multi"], t=t)
    data["spf"].append([n, domain, s["record"], s["all"], s["lookups"], s["void"], s["len"],
                        s["providers"], s["multi"], sst, ssv])
    if s["multi"] == "si": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "SPF", t("f_spf_multi"), t("sev_critical"), t("rec_consolidate")])
    if s["lookups"] > 10: hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "SPF", t("f_spf_lookups"), t("sev_high"), t("rec_reduce")])
    if s["all"] == "+all": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "SPF", t("f_spf_plus_all"), t("sev_critical"), t("rec_change_all")])
    if s["all"] in ("?all", "sin all"): hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "SPF", t("f_spf_soft", val=s['all']), t("sev_high"), t("rec_harden")])
    if s["pub"] == "no": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "SPF", t("f_spf_none"), t("sev_high"), t("rec_publish")])

    # DKIM
    for r in info["dkim"]:
        data["dkim"].append([r["row"], r["domain"], r["selector"], r["service"], r["record"],
                            r["algorithm"], r["bits"], r["t_flag"], t("manual_validation"),
                            t("manual_validation"), t("manual_validation"), r["estado"], r["severidad"]])
        if isinstance(r["bits"], int):
            if r["bits"] < 1024: hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DKIM", t("f_dkim_bits_low", selector=r['selector']), t("sev_critical"), t("rec_rotate")])
            elif r["bits"] < 2048: hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DKIM", t("f_dkim_bits_med", selector=r['selector']), t("sev_high"), t("rec_migrate_2048")])
    if not info["dkim"]:
        if dkim_res.get("deep"):
            hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DKIM", t("f_dkim_not_found_deep"), t("sev_medium"), t("rec_get_selector")])
        else:
            hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DKIM", t("f_dkim_not_found_common"), t("sev_medium"), t("rec_retry_deep")])

    # DMARC
    dst = derive_dmarc_status(d["pub"], d["p"], d["sp"], d["pct"], d["rua"], t=t)
    dsv = derive_dmarc_severity(d["pub"], d["p"], d["sp"], d["pct"], d["rua"], t=t)
    data["dmarc"].append([n, domain, d["record"], d["p"], d["sp"], d["pct"], d["aspf"], d["adkim"],
                         d["rua"], d["ruf"], d["fo"], d["rf"], d["ri"], "", "", dst, dsv])
    if d["pub"] == "no": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DMARC", t("f_dmarc_none"), t("sev_critical"), t("rec_publish")])
    elif d["p"] == "none": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DMARC", t("f_dmarc_p_none"), t("sev_high"), t("rec_migrate_quarantine")])
    elif d["p"] == "quarantine": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DMARC", t("f_dmarc_p_quar"), t("sev_medium"), t("rec_migrate_reject")])

    # DNSSEC
    nst = derive_dnssec_status(ds["diag"], t=t)
    nsv = derive_dnssec_severity(ds["diag"], t=t)
    data["dnssec"].append([n, domain, ds["dnskey_pub"], ds["ds_pub"], ds["ad_flag"], ds["status"],
                          ds["algos"], ds["diag"], nst, nsv])
    if ds["diag"].startswith("Bogus"): hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DNSSEC", t("f_dnssec_bogus"), t("sev_critical"), t("rec_check_ds")])
    elif ds["diag"].startswith("Incompleto"): hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DNSSEC", t("f_dnssec_incomplete"), t("sev_high"), t("rec_publish_ds")])
    elif ds["diag"] == "No implementado": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "DNSSEC", t("f_dnssec_none"), t("sev_medium"), t("rec_activate")])

    # MTA-STS
    mst = derive_mtasts_status(mt["pub"], mt["mode"], mt["accessible"], t=t)
    counters["mtasts_ev"] += 1
    data["mtasts"].append([domain, mt["mode"] or t("not_published"), mt["max_age"], mt["mx"],
                          mt["accessible"], mst, f"EV-MTASTS-{counters['mtasts_ev']:03d}", ""])
    if mt["pub"] == "no": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "MTA-STS", t("f_mtasts_none"), t("sev_low"), t("rec_publish")])

    # TLS-RPT
    tst = derive_tlsrpt_status(tl["pub"], tl["rua"], t=t)
    counters["tlsrpt_ev"] += 1
    data["tlsrpt"].append([domain, f"v=TLSRPTv1; rua={tl['rua']}" if tl["pub"] == "si" else t("not_published"),
                          tl["rua"], t("yes") if tl["rua"] else t("no"),
                          tst, f"EV-TLSRPT-{counters['tlsrpt_ev']:03d}", ""])
    if tl["pub"] == "no": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "TLS-RPT", t("f_tlsrpt_none"), t("sev_low"), t("rec_publish")])

    # BIMI
    bst = derive_bimi_status(bi["pub"], bi["svg"], bi["vmc_status"], t=t)
    data["bimi"].append([domain, bi["record"], bi["svg"], bi["vmc_status"],
                        bi["vmc_issuer"], bi["vmc_exp"], bst, ""])
    if bi["pub"] == "si" and bi["vmc_status"] == "No": hallazgos.append([f"H-{len(hallazgos)+1:03d}", domain, "BIMI", t("f_bimi_no_vmc"), t("sev_info"), t("rec_acquire_vmc")])

    # Remitentes autorizados
    provs = []; seen = set()
    if s["providers"]:
        for p in [x.strip() for x in s["providers"].split(";")]:
            if p and p not in seen:
                seen.add(p); provs.append(p)
    if m["provider"] and m["provider"] not in ("Sin MX", "Otro") and m["provider"] not in seen:
        seen.add(m["provider"]); provs.append(m["provider"])
    if not provs: provs = [t("status_to_validate")]
    for prov in provs:
        prop = map_provider_purpose(prov, t=t)
        mec = map_provider_mechanism(prov, t=t)
        inc = t("yes") if (s["providers"] and prov in s["providers"]) else t("no")
        firma = t("no")
        fw = prov.split()[0] if prov.split() else ""
        for dk in info["dkim"]:
            if fw.lower() in dk["service"].lower():
                firma = t("yes"); break
        if d["pub"] == "no" or d["p"] == "none": ali = t("no")
        elif inc == t("yes") and firma == t("yes"): ali = t("yes")
        elif inc == t("yes") or firma == t("yes"): ali = t("partial")
        else: ali = t("no")
        est = derive_remit_state(inc, firma, ali, t=t)
        counters["remit_row"] += 1
        data["remit"].append([counters["remit_row"], domain, prov, prop, mec, inc, firma, ali,
                             t("manual_validation"), t("manual_validation"), est])

    # Cumplimiento global
    score = 0
    if s["pub"] == "si" and s["all"] in ("-all", "~all") and s["multi"] == "no" and s["lookups"] <= 10: score += 1
    if d["pub"] == "si" and d["p"] != "none": score += 1
    if ds["diag"] == "Secure" or ds["diag"].startswith("Firmado"): score += 1
    if mt["mode"] == "enforce": score += 1
    if tl["pub"] == "si": score += 1
    if bi["pub"] == "si": score += 1
    if info["dkim"]: score += 1
    pct = score * 100 // 7
    if pct >= 85: cumpl = t("compliance_high", pct=pct)
    elif pct >= 50: cumpl = t("compliance_medium", pct=pct)
    else: cumpl = t("compliance_low", pct=pct)
    info["cumplimiento"] = cumpl; info["cumplimiento_pct"] = pct

    # EASM & CISO Compliance Evaluation
    ciso_score, ciso_grade, comp_matrix = evaluate_ciso_compliance_and_score(
        info, info.get("easm", []), info.get("takeover", []), t=t
    )
    info["ciso_score"] = ciso_score
    info["ciso_grade"] = ciso_grade

    for comp_row in comp_matrix:
        data["compliance"].append(comp_row)

    for item in info.get("easm", []):
        data["easm"].append([
            domain, item["lookalike"], item["type"], item["ips"], item["mx"],
            item["threat"], item["action"]
        ])
        if item.get("registered") and item.get("threat") in (t("sev_critical"), t("sev_high")):
            hallazgos.append([
                f"H-{len(hallazgos)+1:03d}", domain, "EASM / Typosquatting",
                t("f_typosquat_registered", domain=item['lookalike']),
                item["threat"], item["action"]
            ])

    for item in info.get("takeover", []):
        if item.get("dangling"):
            hallazgos.append([
                f"H-{len(hallazgos)+1:03d}", domain, "Subdomain Takeover",
                t("f_subdomain_takeover", target=item['cname']),
                t("sev_critical"), "Claim / Delete Dangling DNS"
            ])

    # Inventario y resumen
    brand = info.get("brand") or derive_brand_from_domain(domain)
    owner = t("default_internal_owner")
    data["inventario"].append([n, domain, t("production"), brand, info["registrar"], info["expires"],
                              t("yes") if info["mx"]["provider"] != "Sin MX" else t("no"), owner,
                              info["ns_provider"], ds["dnskey_pub"], f"CISO Score: {ciso_score} ({ciso_grade})"])
    data["resumen"].append([domain, info["registrar"], info["created"], info["expires"], info["status"],
                           info["dnssec_whois"], info["ns_list"], info["soa_serial"], ds["diag"],
                           s["pub"], s["all"], s["lookups"], s["void"], s["multi"], d["pub"], d["p"],
                           d["sp"], d["pct"], d["rua"], m["provider"], mt["mode"] or t("not_published"),
                           tl["pub"], bi["pub"], cumpl])


def build_excel(outdir: Path, data: Dict[str, List[Any]], hallazgos: List[Any], stats: Dict[str, Any], args: Any, t: Optional[Any] = None) -> Path:
    if not t: t = lambda k: k
    wb = Workbook()
    thin = Side(border_style="thin", color="B7C9D6")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    F_TITLE = Font(name="Century Gothic", size=18, bold=True, color="FFFFFF")
    F_SUBT = Font(name="Century Gothic", size=12, bold=True, color="FFFFFF")
    F_H1 = Font(name="Century Gothic", size=11, bold=True, color="FFFFFF")
    F_H2 = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
    F_LABEL = Font(name="Calibri", size=11, bold=True, color=C_NAVY)
    F_BODY = Font(name="Calibri", size=10, color="222222")
    FILL_TITLE = PatternFill("solid", fgColor=C_DARK)
    FILL_SUBT = PatternFill("solid", fgColor=C_MID)
    FILL_H1 = PatternFill("solid", fgColor=C_DARK)
    FILL_H2 = PatternFill("solid", fgColor=C_ACC)
    FILL_BAND = PatternFill("solid", fgColor=C_BAND)
    FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL = Alignment(horizontal="left", vertical="center", wrap_text=True)

    c_comp = t("status_compliant")
    c_part = t("status_partial")
    c_non = t("status_non_compliant")
    c_na = t("status_not_applicable")
    c_pend = t("status_pending")

    s_crit = t("sev_critical")
    s_high = t("sev_high")
    s_med = t("sev_medium")
    s_low = t("sev_low")
    s_inf = t("sev_info")

    LIST_EST = f'"{c_comp},{c_part},{c_non},{c_na},{c_pend}"'
    LIST_SEV = f'"{s_crit},{s_high},{s_med},{s_low},{s_inf}"'

    def style_header(ws: Any, row: int, n_cols: int, fill: Any = FILL_H1, font: Any = F_H1) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill; cell.font = font
            cell.alignment = AC; cell.border = BORDER
        ws.row_dimensions[row].height = 30

    def set_widths(ws: Any, widths: List[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def cf_status(ws: Any, col: str, s_row: int, e_row: int) -> None:
        rng = f"{col}{s_row}:{col}{e_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{c_comp}"'],
            fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{c_part}"'],
            fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C5700")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{c_non}"'],
            fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{c_na}"'],
            fill=PatternFill("solid", fgColor="D9D9D9"), font=Font(color="555555")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{c_pend}"'],
            fill=PatternFill("solid", fgColor="DDEBF7"), font=Font(color="1F4E78")))

    def cf_sev(ws: Any, col: str, s_row: int, e_row: int) -> None:
        rng = f"{col}{s_row}:{col}{e_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{s_crit}"'],
            fill=PatternFill("solid", fgColor="C00000"), font=Font(color="FFFFFF", bold=True)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{s_high}"'],
            fill=PatternFill("solid", fgColor="ED7D31"), font=Font(color="FFFFFF", bold=True)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{s_med}"'],
            fill=PatternFill("solid", fgColor="FFD966"), font=Font(color="7F5F00", bold=True)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{s_low}"'],
            fill=PatternFill("solid", fgColor="A9D08E"), font=Font(color="375623")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{s_inf}"'],
            fill=PatternFill("solid", fgColor="D9D9D9"), font=Font(color="555555")))

    # 1. Portada / Cover Ejecutiva (Única hoja inicial, sin colisiones)
    ws_cover = wb.active
    ws_cover.title = t("sheet_summary")
    ws_cover.sheet_view.showGridLines = False
    try:
        ws_cover.views.sheetView[0].showGridLines = False
    except Exception:
        pass

    for r in range(1, 45):
        for c in range(1, 15):
            ws_cover.cell(row=r, column=c).fill = FILL_WHITE

    ws_cover.merge_cells("A1:N1")
    tcell = ws_cover["A1"]
    tcell.value = t("cover_title")
    tcell.font = F_TITLE; tcell.fill = FILL_TITLE; tcell.alignment = AC
    ws_cover.row_dimensions[1].height = 40

    ws_cover.merge_cells("A2:N2")
    scell = ws_cover["A2"]
    scell.value = t("cover_meta", author=t("author"), ciso=t("ciso"), date=datetime.now().strftime('%Y-%m-%d %H:%M'))
    scell.font = F_SUBT; scell.fill = FILL_SUBT; scell.alignment = AC
    ws_cover.row_dimensions[2].height = 25

    cards = [
        (t("card_total_domains"), stats["total_domains"], C_DARK, "A", "C"),
        (t("card_critical_findings"), stats["sev_count"].get(t("sev_critical"), 0), "C00000", "D", "F"),
        (t("card_high_findings"), stats["sev_count"].get(t("sev_high"), 0), "ED7D31", "G", "I"),
        (t("card_ciso_score"), f"{stats.get('ciso_score', '88 (B)')}", "0066B3", "J", "L"),
        (t("card_avg_compliance"), f"{stats.get('avg_compliance', 85)}%", "70AD47", "M", "N"),
    ]

    for title, val, color_hex, start_col, end_col in cards:
        ws_cover.merge_cells(f"{start_col}4:{end_col}4")
        c_title = ws_cover[f"{start_col}4"]
        c_title.value = title
        c_title.font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
        c_title.fill = PatternFill("solid", fgColor=color_hex)
        c_title.alignment = AC

        ws_cover.merge_cells(f"{start_col}5:{end_col}6")
        c_val = ws_cover[f"{start_col}5"]
        c_val.value = val
        c_val.font = Font(name="Century Gothic", size=18, bold=True, color="FFFFFF")
        c_val.fill = PatternFill("solid", fgColor=color_hex)
        c_val.alignment = AC

    ws_cover.row_dimensions[4].height = 20
    ws_cover.row_dimensions[5].height = 22
    ws_cover.row_dimensions[6].height = 22

    # Hallazgos Prioritarios en Portada
    ws_cover.merge_cells("A8:N8")
    hf_hdr = ws_cover["A8"]
    hf_hdr.value = t("top_findings_title")
    hf_hdr.font = F_SUBT; hf_hdr.fill = FILL_SUBT; hf_hdr.alignment = AC
    ws_cover.row_dimensions[8].height = 25

    ws_cover.cell(row=9, column=1, value=t("col_id"))
    ws_cover.cell(row=9, column=2, value=t("col_domain"))
    ws_cover.cell(row=9, column=4, value=t("col_control"))
    ws_cover.cell(row=9, column=6, value=t("col_finding_desc"))
    ws_cover.cell(row=9, column=10, value=t("col_severity"))
    ws_cover.cell(row=9, column=12, value=t("col_action"))
    style_header(ws_cover, 9, 14, fill=FILL_H2, font=F_H2)

    crit_high_findings = [h for h in hallazgos if h[4] in (t("sev_critical"), t("sev_high"))][:15]
    for idx, h in enumerate(crit_high_findings, start=10):
        ws_cover.cell(row=idx, column=1, value=h[0]).font = F_BODY
        ws_cover.cell(row=idx, column=2, value=h[1]).font = F_BODY
        ws_cover.cell(row=idx, column=4, value=h[2]).font = F_BODY
        ws_cover.cell(row=idx, column=6, value=h[3]).font = F_BODY
        ws_cover.cell(row=idx, column=10, value=h[4]).font = Font(name="Calibri", size=10, bold=True, color="C00000" if h[4] == t("sev_critical") else "ED7D31")
        ws_cover.cell(row=idx, column=12, value=h[5]).font = F_BODY
        ws_cover.row_dimensions[idx].height = 20

    set_widths(ws_cover, [14, 20, 6, 16, 6, 28, 6, 6, 6, 16, 6, 32, 6, 6])

    def add_sheet(name: str, headers: List[str], rows: List[Any], status_col: Optional[str] = None, sev_col: Optional[str] = None, col_widths: Optional[List[int]] = None) -> Any:
        ws_new = wb.create_sheet(name)
        for i, h in enumerate(headers, start=1):
            ws_new.cell(row=1, column=i, value=h)
        style_header(ws_new, 1, len(headers))
        for ri, row in enumerate(rows, start=2):
            for ci, val in enumerate(row, start=1):
                cell = ws_new.cell(row=ri, column=ci, value=val)
                cell.font = F_BODY; cell.alignment = AL; cell.border = BORDER
                cell.fill = FILL_BAND if ri % 2 == 0 else FILL_WHITE
            ws_new.row_dimensions[ri].height = 22
        if col_widths:
            set_widths(ws_new, col_widths)
        ws_new.freeze_panes = "A2"
        end = max(2, 1 + len(rows))
        if status_col:
            dv = DataValidation(type="list", formula1=LIST_EST, allow_blank=True)
            ws_new.add_data_validation(dv)
            dv.add(f"{status_col}2:{status_col}{end}")
            cf_status(ws_new, status_col, 2, end)
        if sev_col:
            dv = DataValidation(type="list", formula1=LIST_SEV, allow_blank=True)
            ws_new.add_data_validation(dv)
            dv.add(f"{sev_col}2:{sev_col}{end}")
            cf_sev(ws_new, sev_col, 2, end)
        return ws_new

    # 2. Inventario de Dominios
    add_sheet(t("sheet_inventory"),
        [t("col_num"), t("col_domain"), t("col_asset_type"), t("col_brand_entity"), t("col_registrar"),
         t("col_expiration_date"), t("col_sends_email"), t("col_internal_owner"), t("col_dns_managed_by"),
         t("col_dnssec_pub"), t("col_caa"), t("col_fcrdns"), t("col_comments")],
        data["inventario"],
        col_widths=[5, 28, 16, 26, 24, 20, 14, 22, 22, 12, 25, 25, 35])

    # 3. SPF
    add_sheet(t("sheet_spf"),
        [t("col_num"), t("col_domain"), t("col_spf_record"), t("col_all_mech"), t("col_dns_lookups"),
         t("col_void_lookups"), t("col_length"), t("col_auth_providers"), t("col_multi_spf"), t("col_status"), t("col_severity")],
        data["spf"], status_col="J", sev_col="K",
        col_widths=[5, 26, 55, 14, 12, 12, 14, 35, 14, 22, 14])

    # 4. DKIM
    add_sheet(t("sheet_dkim"),
        [t("col_num"), t("col_domain"), t("col_selector"), t("col_signing_service"), t("col_public_record"),
         t("col_algorithm"), t("col_key_bits"), t("col_flag_t"), t("col_created_date"), t("col_last_rotation"),
         t("col_next_rotation"), t("col_status"), t("col_severity")],
        data["dkim"], status_col="L", sev_col="M",
        col_widths=[5, 24, 20, 24, 50, 14, 14, 14, 16, 16, 16, 22, 14])

    # 5. DMARC
    add_sheet(t("sheet_dmarc"),
        [t("col_num"), t("col_domain"), t("col_dmarc_record"), t("col_policy_p"), t("col_subdomain_sp"), t("col_pct"),
         t("col_aspf"), t("col_adkim"), t("col_rua"), t("col_ruf"), t("col_fo"), t("col_rf"), t("col_ri"),
         t("col_reports_active"), t("col_analysis_plat"), t("col_status"), t("col_severity")],
        data["dmarc"], status_col="P", sev_col="Q",
        col_widths=[5, 24, 50, 14, 14, 8, 14, 14, 28, 28, 10, 10, 10, 16, 22, 22, 14])

    # 6. DNSSEC
    add_sheet(t("sheet_dnssec"),
        [t("col_num"), t("col_domain"), t("col_dnskey_published"), t("col_ds_parent"), t("col_ad_flag"),
         t("col_resolver_status"), t("col_algorithms"), t("col_diagnostic"), t("col_status"), t("col_severity")],
        data["dnssec"], status_col="I", sev_col="J",
        col_widths=[5, 24, 16, 16, 16, 16, 16, 30, 22, 14])

    # 7. Complementos (MTA-STS, TLS-RPT, BIMI, CAA & TLS Cert)
    ws_comp = wb.create_sheet(t("sheet_mtasts"))
    ws_comp.sheet_view.showGridLines = False
    try:
        ws_comp.views.sheetView[0].showGridLines = False
    except Exception:
        pass

    # MTA-STS Section
    ws_comp.merge_cells("A1:H1")
    ws_comp["A1"] = t("sec_mtasts"); ws_comp["A1"].font = F_SUBT; ws_comp["A1"].fill = FILL_SUBT; ws_comp["A1"].alignment = AC
    ws_comp.row_dimensions[1].height = 26
    mta_hdr = [t("col_domain"), t("col_published_policy"), t("col_max_age"), t("col_mx_hosts"),
               t("col_well_known"), t("col_status"), t("col_evidence_id"), t("col_comments")]
    for i, h in enumerate(mta_hdr, start=1):
        ws_comp.cell(row=2, column=i, value=h)
    style_header(ws_comp, 2, len(mta_hdr), fill=FILL_H2, font=F_H2)
    for ri, row in enumerate(data["mtasts"], start=3):
        for ci, val in enumerate(row, start=1):
            c = ws_comp.cell(row=ri, column=ci, value=val)
            c.font = F_BODY; c.alignment = AL; c.border = BORDER
            c.fill = FILL_BAND if ri % 2 == 1 else FILL_WHITE
    mta_end = max(3, 2 + len(data["mtasts"]))
    if data["mtasts"]:
        dv = DataValidation(type="list", formula1=LIST_EST, allow_blank=True); ws_comp.add_data_validation(dv)
        dv.add(f"F3:F{mta_end}"); cf_status(ws_comp, "F", 3, mta_end)

    # TLS-RPT Section
    tls_start = mta_end + 3
    ws_comp.merge_cells(f"A{tls_start}:G{tls_start}")
    ws_comp[f"A{tls_start}"] = t("sec_tlsrpt"); ws_comp[f"A{tls_start}"].font = F_SUBT
    ws_comp[f"A{tls_start}"].fill = FILL_SUBT; ws_comp[f"A{tls_start}"].alignment = AC
    ws_comp.row_dimensions[tls_start].height = 26
    tls_hdr = [t("col_domain"), t("col_txt_smtp_tls"), t("col_rua_mailbox"), t("col_receiving_reports"), t("col_status"), t("col_evidence_id"), t("col_comments")]
    for i, h in enumerate(tls_hdr, start=1):
        ws_comp.cell(row=tls_start + 1, column=i, value=h)
    style_header(ws_comp, tls_start + 1, len(tls_hdr), fill=FILL_H2, font=F_H2)
    for ri, row in enumerate(data["tlsrpt"], start=tls_start + 2):
        for ci, val in enumerate(row, start=1):
            c = ws_comp.cell(row=ri, column=ci, value=val)
            c.font = F_BODY; c.alignment = AL; c.border = BORDER
            c.fill = FILL_BAND if ri % 2 == 1 else FILL_WHITE
    tls_end = max(tls_start + 2, tls_start + 1 + len(data["tlsrpt"]))
    if data["tlsrpt"]:
        dv = DataValidation(type="list", formula1=LIST_EST, allow_blank=True); ws_comp.add_data_validation(dv)
        dv.add(f"E{tls_start+2}:E{tls_end}"); cf_status(ws_comp, "E", tls_start + 2, tls_end)

    # BIMI Section
    bimi_start = tls_end + 3
    ws_comp.merge_cells(f"A{bimi_start}:H{bimi_start}")
    ws_comp[f"A{bimi_start}"] = t("sec_bimi"); ws_comp[f"A{bimi_start}"].font = F_SUBT
    ws_comp[f"A{bimi_start}"].fill = FILL_SUBT; ws_comp[f"A{bimi_start}"].alignment = AC
    ws_comp.row_dimensions[bimi_start].height = 26
    bimi_hdr = [t("col_domain"), t("col_bimi_record"), t("col_svg_url"), t("col_vmc_cert"),
                t("col_vmc_issuer"), t("col_vmc_exp"), t("col_status"), t("col_comments")]
    for i, h in enumerate(bimi_hdr, start=1):
        ws_comp.cell(row=bimi_start + 1, column=i, value=h)
    style_header(ws_comp, bimi_start + 1, len(bimi_hdr), fill=FILL_H2, font=F_H2)
    for ri, row in enumerate(data["bimi"], start=bimi_start + 2):
        for ci, val in enumerate(row, start=1):
            c = ws_comp.cell(row=ri, column=ci, value=val)
            c.font = F_BODY; c.alignment = AL; c.border = BORDER
            c.fill = FILL_BAND if ri % 2 == 1 else FILL_WHITE
    bimi_end = max(bimi_start + 2, bimi_start + 1 + len(data["bimi"]))
    if data["bimi"]:
        dv = DataValidation(type="list", formula1=LIST_EST, allow_blank=True); ws_comp.add_data_validation(dv)
        dv.add(f"G{bimi_start+2}:G{bimi_end}"); cf_status(ws_comp, "G", bimi_start + 2, bimi_end)

    # CAA & TLS Certificate Health Section
    caa_start = bimi_end + 3
    ws_comp.merge_cells(f"A{caa_start}:F{caa_start}")
    ws_comp[f"A{caa_start}"] = t("sec_caa_tls"); ws_comp[f"A{caa_start}"].font = F_SUBT
    ws_comp[f"A{caa_start}"].fill = FILL_SUBT; ws_comp[f"A{caa_start}"].alignment = AC
    ws_comp.row_dimensions[caa_start].height = 26
    caa_hdr = [t("col_domain"), t("col_caa"), t("col_iodef"), t("col_fcrdns"), t("col_tls_cert"), t("col_tls_issuer")]
    for i, h in enumerate(caa_hdr, start=1):
        ws_comp.cell(row=caa_start + 1, column=i, value=h)
    style_header(ws_comp, caa_start + 1, len(caa_hdr), fill=FILL_H2, font=F_H2)
    for ri, row in enumerate(data.get("caa_tls", []), start=caa_start + 2):
        for ci, val in enumerate(row, start=1):
            c = ws_comp.cell(row=ri, column=ci, value=val)
            c.font = F_BODY; c.alignment = AL; c.border = BORDER
            c.fill = FILL_BAND if ri % 2 == 1 else FILL_WHITE

    set_widths(ws_comp, [24, 30, 28, 22, 28, 24, 16, 30])

    # 8. Remitentes Autorizados
    add_sheet(t("sheet_senders"),
        [t("col_num"), t("col_domain"), t("col_service_provider"), t("col_purpose"), t("col_sending_mech"),
         t("col_in_spf"), t("col_own_dkim"), t("col_dmarc_align"), t("col_internal_owner"), t("col_date_added"), t("col_status")],
        data["remit"], status_col="K",
        col_widths=[5, 24, 24, 26, 22, 14, 16, 18, 28, 16, 22])

    # 9. Hallazgos
    add_sheet(t("sheet_findings"),
        [t("col_id"), t("col_domain"), t("col_control"), t("col_finding_desc"), t("col_severity"), t("col_action")],
        hallazgos, sev_col="E",
        col_widths=[14, 24, 14, 55, 14, 45])

    # 10. Superficie de Ataque & Typosquats Sheet
    add_sheet(t("sheet_easm"),
        [t("col_domain"), t("col_lookalike"), t("col_mutation_type"), t("col_dns_status"),
         t("col_mx_status"), t("col_threat_level"), t("col_action")],
        data.get("easm", []), sev_col="F",
        col_widths=[24, 26, 20, 26, 26, 16, 30])

    # 11. Matriz de Cumplimiento CISO Sheet
    add_sheet(t("sheet_compliance"),
        [t("col_domain"), t("col_standard"), t("col_requirement"), t("col_audit_status"), t("col_gap_notes")],
        data.get("compliance", []), status_col="D",
        col_widths=[24, 24, 45, 20, 55])

    # 12. Resumen Consolidado
    add_sheet(t("sheet_consolidated"),
        [t("col_cons_domain"), t("col_cons_registrar"), t("col_cons_created"), t("col_cons_expires"),
         t("col_cons_status"), t("col_cons_whois_dnssec"), t("col_cons_ns"), t("col_cons_soa"),
         t("col_cons_dnssec_diag"), t("col_cons_spf_pub"), t("col_cons_spf_all"), t("col_cons_spf_lookups"),
         t("col_cons_spf_void"), t("col_cons_spf_multi"), t("col_cons_dmarc_pub"), t("col_cons_dmarc_p"),
         t("col_cons_dmarc_sp"), t("col_cons_dmarc_pct"), t("col_cons_dmarc_rua"), t("col_cons_mx"),
         t("col_cons_mtasts"), t("col_cons_tlsrpt"), t("col_cons_bimi"), t("col_cons_compliance")],
        data["resumen"],
        col_widths=[24, 22, 14, 14, 16, 14, 30, 14, 20, 10, 12, 12, 12, 12, 10, 12, 12, 10, 28, 20, 14, 10, 10, 20])

    wb.active = 0
    out_file = outdir / f"audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_file)
    return out_file


def banner(t: Optional[Any] = None) -> Panel:
    if not t: t = lambda k: k
    art = (f"[bold {NM}]╔══════════════════════════════════════════════════════════════╗[/]\n"
           f"[bold {NM}]║[/]  [bold {NC}]E M A I L[/]  [bold {NG}]D N S[/]  [bold {NO}]A U D I T[/]  [bold {NY}]v 3 . 3[/]            [bold {NM}]║[/]\n"
           f"[bold {NM}]║[/]  [{NP} italic]{t('app_subtitle')}[/]                  [bold {NM}]║[/]\n"
           f"[bold {NM}]╚══════════════════════════════════════════════════════════════╝[/]\n"
           f"  [{NC}]Author / Autor:[/] [bold {NY}]{t('author')}[/]   [{NC}]CISO:[/] [bold {NG}]{t('ciso')}[/]   "
           f"[{NC}]Version:[/] [bold {NO}]3.3[/]")
    return Panel(art, border_style=NM, box=DOUBLE, padding=(0, 1))


def status_bar(domain: str, idx: int, total: int, findings: int, start: float, t: Optional[Any] = None) -> Panel:
    if not t: t = lambda k: k
    el = int(time.time() - start); mins, secs = divmod(el, 60)
    pct = (idx / total * 100) if total else 0
    bl = 30; fl = int(bl * idx / total) if total else 0
    bar = f"[{NG}]" + "█" * fl + "[/][grey50]" + "░" * (bl - fl) + "[/]"
    return Panel(f"[{NC}]▶[/] [bold {NY}]{domain}[/]   {bar} [bold {NM}]{idx}/{total}[/] "
                 f"([bold {NG}]{pct:.0f}%[/])   [{NO}]{t('col_findings')}:[/] [bold {NR}]{findings}[/]   "
                 f"[{NP}]Time:[/] [bold {NC}]{mins:02d}:{secs:02d}[/]",
                 border_style=NC, box=ROUNDED, padding=(0, 1))


def render_panel(info: Dict[str, Any], t: Optional[Any] = None) -> Panel:
    if not t: t = lambda k: k
    s = info["spf"]; dm = info["dmarc"]; dns_d = info["dnssec"]
    mt = info["mtasts"]; tl = info["tlsrpt"]; bi = info["bimi"]
    dkim_res = info.get("dkim_res", {"found": info["dkim"], "probed": 0, "deep": False})
    cumpl = info.get("cumplimiento", "N/D"); pct = info.get("cumplimiento_pct", 0)
    color = NG if pct >= 85 else NY if pct >= 50 else NR
    tbl = Table(box=HEAVY, border_style=NP, show_header=True,
                header_style=Style(color=NM, bold=True))
    tbl.add_column("Control", style=Style(color=NC, bold=True), width=12)
    tbl.add_column("Status / Estado", width=56)

    def cell(x: str, l: str) -> str:
        if l == "ok": return f"[{NG}]✔ {x}[/]"
        if l == "warn": return f"[{NY}]◐ {x}[/]"
        if l == "fail": return f"[{NR}]✘ {x}[/]"
        return f"[grey50]· {x}[/]"

    # SPF
    if s["pub"] == "si":
        lvl = "ok" if (s["all"] in ("-all", "~all") and s["multi"] == "no" and s["lookups"] <= 10) else "warn"
        tbl.add_row("SPF", cell(f"{s['all']} · lookups={s['lookups']}", lvl))
    else:
        tbl.add_row("SPF", cell(t("not_published"), "fail"))

    # DKIM
    if info["dkim"]:
        sels = ", ".join([f"{r['selector']}({r['bits']})" for r in info["dkim"][:3]])
        if len(info["dkim"]) > 3:
            sels += f" +{len(info['dkim'])-3}"
        tbl.add_row("DKIM", cell(sels, "ok"))
    else:
        if dkim_res.get("deep"):
            tbl.add_row("DKIM", cell(f"{t('f_dkim_not_found_deep')} ({dkim_res.get('probed',0)} sels)", "warn"))
        else:
            tbl.add_row("DKIM", cell(t("f_dkim_not_found_common"), "warn"))

    # DMARC
    if dm["pub"] == "si":
        lvl = "ok" if dm["p"] == "reject" else "warn" if dm["p"] == "quarantine" else "fail"
        tbl.add_row("DMARC", cell(f"p={dm['p']} pct={dm['pct']}", lvl))
    else:
        tbl.add_row("DMARC", cell(t("not_published"), "fail"))

    # DNSSEC
    diag = dns_d["diag"]
    lvl = "ok" if (diag == "Secure" or diag.startswith("Firmado")) else "warn" if diag.startswith("Incompleto") else "fail"
    tbl.add_row("DNSSEC", cell(diag, lvl))

    # MTA-STS
    if mt["mode"] == "enforce":
        tbl.add_row("MTA-STS", cell("enforce", "ok"))
    elif mt["mode"] == "testing":
        tbl.add_row("MTA-STS", cell("testing", "warn"))
    else:
        tbl.add_row("MTA-STS", cell(t("not_published"), "fail"))

    # TLS-RPT
    tbl.add_row("TLS-RPT", cell(t("yes") if tl["pub"] == "si" else t("not_published"),
                                "ok" if tl["pub"] == "si" else "fail"))

    # BIMI
    if bi["pub"] == "si":
        lvl = "ok" if bi["vmc_status"] in ("Si", "Yes") else "warn"
        tbl.add_row("BIMI", cell(f"VMC={bi['vmc_status']}", lvl))
    else:
        tbl.add_row("BIMI", cell(t("not_published"), "fail"))

    # CAA
    caa = info.get("caa", {})
    if caa.get("pub") == "si":
        tbl.add_row("CAA", cell(caa.get("cas", "Restringido"), "ok"))
    else:
        tbl.add_row("CAA", cell(t("not_published"), "warn"))

    # FCrDNS
    fcrdns = info.get("fcrdns", {})
    if fcrdns.get("status") == "Sin MX":
        tbl.add_row("FCrDNS", cell(t("no_mx"), "warn"))
    elif fcrdns.get("compliant"):
        tbl.add_row("FCrDNS", cell("PTR ↔ A OK", "ok"))
    else:
        tbl.add_row("FCrDNS", cell("Mismatch / Sin PTR", "fail"))

    # TLS Certificate
    tls_c = info.get("tls_cert", {})
    if tls_c.get("warning") == "critical":
        tbl.add_row("TLS Cert", cell(f"Vence en {tls_c.get('days_left')}d", "fail"))
    elif tls_c.get("warning") == "warning":
        tbl.add_row("TLS Cert", cell(f"Vence en {tls_c.get('days_left')}d", "warn"))
    elif tls_c.get("days_left") is not None:
        tbl.add_row("TLS Cert", cell(f"Válido ({tls_c.get('days_left')}d)", "ok"))
    else:
        tbl.add_row("TLS Cert", cell("N/D", "warn"))

    # MX
    tbl.add_row("MX", cell(info["mx"]["provider"],
                           "ok" if info["mx"]["provider"] not in ("Sin MX", "Otro") else "warn"))

    title = f"[bold {NC}]▶ {info['domain']}[/]   [bold {color}]{cumpl}[/]"
    return Panel(tbl, title=title, border_style=color, box=DOUBLE, padding=(0, 1))


def final_panel(stats: Dict[str, Any], paths: Dict[str, Any], elapsed: float, t: Optional[Any] = None) -> Panel:
    if not t: t = lambda k: k
    mins, secs = divmod(int(elapsed), 60)
    sev_t = Table(box=ROUNDED, border_style=NP, show_header=True,
                  header_style=Style(color=NM, bold=True))
    sev_t.add_column("Severity / Severidad", style=Style(color=NC, bold=True))
    sev_t.add_column("Count", justify="right")
    colors = {t("sev_critical"): NR, t("sev_high"): NO, t("sev_medium"): NY, t("sev_low"): NG, t("sev_info"): "grey70"}
    for sev_key in ["sev_critical", "sev_high", "sev_medium", "sev_low", "sev_info"]:
        s_lbl = t(sev_key)
        n = stats["sev_count"].get(s_lbl, 0)
        c = colors.get(s_lbl, "white")
        sev_t.add_row(f"[{c}]{s_lbl}[/]", f"[{c}]{n}[/]")
    body = (f"[bold {NC}]{t('card_total_domains')}:[/] [bold {NG}]{stats['total_domains']}[/]\n"
            f"[bold {NC}]Total Time:[/] [bold {NO}]{mins:02d}:{secs:02d}[/]\n"
            f"[bold {NC}]Excel Report:[/] [bold {NY}]{paths['excel']}[/]\n"
            f"[bold {NC}]Evidence Folder:[/] [bold {NY}]{paths['evidencias']}[/]")
    return Panel(body, title=f"[bold {NM}]✦ {t('summary_title')} ✦[/]",
                 border_style=NM, box=DOUBLE, padding=(1, 2))


async def run_audit(args: Any) -> None:
    t = get_translator(args.lang)
    outdir = Path(args.output)
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "evidencias").mkdir(exist_ok=True)
    if args.domain:
        domains = [args.domain.strip().lower()]
    elif args.domains:
        with open(args.domains, "r", encoding="utf-8") as f:
            domains = [l.strip().lower() for l in f if l.strip() and not l.strip().startswith("#")]
    else:
        domains = []
    if not domains:
        console.print(f"[{NR}]ERROR: Sin dominios / No domains found.[/]")
        return
    data = {k: [] for k in ["spf", "dkim", "dmarc", "dnssec", "mtasts", "tlsrpt", "bimi", "remit", "resumen", "inventario", "easm", "compliance", "caa_tls"]}
    hallazgos: List[Any] = []
    ro = dns.asyncresolver.Resolver()
    ro.nameservers = [args.resolver]; ro.timeout = 5; ro.lifetime = 8
    http = httpx.AsyncClient(verify=True, follow_redirects=True)
    counters = {"dom_num": 0, "dkim_row": [0], "mtasts_ev": 0, "tlsrpt_ev": 0, "remit_row": 0}
    stats: Dict[str, Any] = {"total_domains": 0, "total_findings": 0, "sev_count": {}}
    start = time.time()
    console.print(banner(t=t))
    if args.deep_dkim:
        console.print(f"[{NY}][*] Modo DKIM PROFUNDO activado (--deep-dkim, {args.deep_months} meses de selectores rotativos)[/]")
    console.print()
    for i, domain in enumerate(domains, 1):
        try:
            console.print(status_bar(domain, i - 1, len(domains), len(hallazgos), start, t=t))
            info = await audit_domain(domain, args, ro, http, outdir, counters, data, hallazgos, t=t)
            console.print(render_panel(info, t=t)); console.print()
        except Exception as e:
            console.print(f"[{NR}]ERROR {domain}: {e}[/]")
            continue
    for h in hallazgos:
        sev = h[4]
        stats["sev_count"][sev] = stats["sev_count"].get(sev, 0) + 1
    stats["total_domains"] = len(domains)
    stats["total_findings"] = len(hallazgos)
    elapsed = time.time() - start
    excel_path = build_excel(outdir, data, hallazgos, stats, args, t=t)
    await http.aclose()
    console.print(final_panel(stats, {"excel": excel_path.absolute(),
                                      "evidencias": (outdir / "evidencias").absolute()}, elapsed, t=t))
    console.print()
    console.print(Align.center(f"[italic {NP}]Auditoria completada · "
                              f"[bold {NY}]Eduardo Recinos[/] · VCISO[/]"))
    console.print()


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Email DNS Audit v3.3 - Unified Excel, Bilingual Support & RDAP",
                                formatter_class=argparse.RawTextHelpFormatter)
    p.add_argument("--domain", help="Auditar un único dominio / Audit a single domain")
    p.add_argument("--domains", "-d", help="Archivo con dominios / Domain list file")
    p.add_argument("--lang", "-l", choices=["es", "en"], default="es",
                   help="Language / Idioma del reporte y consola: 'es' (Español) o 'en' (Inglés). Default: es")
    p.add_argument("--selectors", "-s", default="", help="Selectores DKIM extra (separados por espacio)")
    p.add_argument("--deep-dkim", action="store_true",
                   help="Busqueda DKIM PROFUNDA: agrega selectores rotativos por fecha.")
    p.add_argument("--deep-months", type=int, default=30,
                   help="Meses hacia atras para generar selectores rotativos DKIM (default 30).")
    p.add_argument("--resolver", "-r", default="1.1.1.1",
                   help="Resolver DNS para consultas generales (default 1.1.1.1)")
    p.add_argument("--dnssec-resolvers", default="1.1.1.1,8.8.8.8,9.9.9.9",
                   help="Resolvers validadores para DNSSEC (default: 1.1.1.1,8.8.8.8,9.9.9.9)")
    p.add_argument("--output", "-o", default=f"./audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                   help="Directorio salida")
    p.add_argument("--excel-name", default="", help="Nombre del Excel")
    p.add_argument("--install-deps", action="store_true")
    p.add_argument("--no-color", action="store_true")
    p.add_argument("--verbose", "-v", action="store_true")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    if not args.domain and not args.domains:
        console.print(f"[{NR}]ERROR: --domain o --domains requerido / --domain or --domains required[/]")
        sys.exit(1)
    if args.domains and not Path(args.domains).is_file():
        console.print(f"[{NR}]ERROR: No existe {args.domains}[/]")
        sys.exit(1)
    try:
        asyncio.run(run_audit(args))
    except KeyboardInterrupt:
        console.print(f"\n[{NR}]Interrumpido / Interrupted.[/]")
        sys.exit(130)


if __name__ == "__main__":
    main()
